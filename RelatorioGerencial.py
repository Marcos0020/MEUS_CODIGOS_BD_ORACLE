# RelatorioGerencial

import cx_Oracle
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime

# Configuração da conexão Oracle
dsn_tns = cx_Oracle.makedsn('192.168.50.7', 2115, service_name='p01')
connection = None
try:
    connection = cx_Oracle.connect(user='INSIDE02', password='A$2$Tvo1LuPLp#U', dsn=dsn_tns)
except cx_Oracle.DatabaseError as e:
    messagebox.showerror("Erro de Conexão", f"Não foi possível conectar ao banco de dados: {e}")

# Função para buscar dados no banco de dados
def buscar_dados():
    try:
        cursor = connection.cursor()

        # Obtendo valores dos filtros
        data_inicio_inclusao = filtro_data_inicio.get()
        data_fim_inclusao = filtro_data_fim.get()
        data_inicio = filtro_data_inicio_data.get()
        data_fim = filtro_data_fim_data.get()
        status_selecionado = status_var.get()
        processo_filtro = filtro_processo.get()

        # Verificando se os campos de data inclusão foram preenchidos corretamente
        if (data_inicio_inclusao and not data_fim_inclusao) or (not data_inicio_inclusao and data_fim_inclusao):
            messagebox.showerror("Erro de Data", "Quando preencher o campo 'Data Inclusão', ambos 'Início' e 'Fim' devem ser preenchidos.")
            return
        
        # Verificando se os campos de data foram preenchidos corretamente
        if (data_inicio and not data_fim) or (not data_inicio and data_fim):
            messagebox.showerror("Erro de Data", "Quando preencher o campo 'Data', ambos 'Início' e 'Fim' devem ser preenchidos.")
            return

        # Validando e convertendo as datas
        try:
            data_inicio_formatada_inclusao = datetime.strptime(data_inicio_inclusao, '%d/%m/%Y').strftime('%Y-%m-%d') if data_inicio_inclusao else None
            data_fim_formatada_inclusao = datetime.strptime(data_fim_inclusao, '%d/%m/%Y').strftime('%Y-%m-%d') if data_fim_inclusao else None
            data_inicio_formatada_data = datetime.strptime(data_inicio, '%d/%m/%Y').strftime('%Y-%m-%d') if data_inicio else None
            data_fim_formatada_data = datetime.strptime(data_fim, '%d/%m/%Y').strftime('%Y-%m-%d') if data_fim else None
        except ValueError as e:
            messagebox.showerror("Erro de Data", "Formato de data inválido. Use o formato DD/MM/YYYY.")
            return

        # Construção da consulta
        query = """
            SELECT CODIGO, TO_CHAR(DATA,'DD/MM/YYYY'), GRUPOPESQUISA, COD_RELACIONAL, NOME, PROCESSO, PAGINA, DIARIO, OBS, TEXTO, STATUS, IMPORTADO, COD_INTERNO, to_char(DATA_INCLUSAO,'DD/MM/YYYY HH24:MM:SS'), ID_ASSUNTO
            FROM CUSTOM.ITDESTAQUELD it
            WHERE 1=1
        """
        params = {}

        # Adiciona o filtro de DATA_INCLUSAO
        if data_inicio_formatada_inclusao and data_fim_formatada_inclusao:
            query += """
                AND TRUNC(DATA_INCLUSAO) >= TO_DATE(:data_inicio_inclusao, 'YYYY-MM-DD') 
                AND TRUNC(DATA_INCLUSAO) < TO_DATE(:data_fim_inclusao, 'YYYY-MM-DD') + INTERVAL '1' DAY
            """
            params['data_inicio_inclusao'] = data_inicio_formatada_inclusao
            params['data_fim_inclusao'] = data_fim_formatada_inclusao

        # Adiciona o filtro de DATA
        if data_inicio_formatada_data and data_fim_formatada_data:
            query += """
                AND TRUNC(DATA) >= TO_DATE(:data_inicio_data, 'YYYY-MM-DD') 
                AND TRUNC(DATA) < TO_DATE(:data_fim_data, 'YYYY-MM-DD') + INTERVAL '1' DAY
            """
            params['data_inicio_data'] = data_inicio_formatada_data
            params['data_fim_data'] = data_fim_formatada_data

        # Filtro por status
        if status_selecionado != "Todos":
            if status_selecionado == "Excluídos":
                query += """
                    AND it.ID_ASSUNTO NOT IN (SELECT ja.ID_ASSUNTO
                                              FROM ldesk.jur_andamento ja 
                                              WHERE ja.id_andamento IS NOT NULL)
                    AND it.status <> 'N'
                """
            else:
                query += " AND STATUS = :status"
                params['status'] = status_selecionado

        # Filtro de "Processo"
        if processo_filtro:
            query += """
                AND (PROCESSO = :processo_filtro 
                     OR PROCESSO LIKE :processo_filtro_wild 
                     OR PROCESSO LIKE :processo_filtro_no_dashes
                     OR TEXTO LIKE :processo_filtro_wild 
                     OR TEXTO LIKE :processo_filtro_no_dashes)
            """
            params['processo_filtro'] = processo_filtro
            params['processo_filtro_wild'] = f"%{processo_filtro}%"
            params['processo_filtro_no_dashes'] = processo_filtro.replace("-", "")

        # Executa a consulta
        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Limpando a tabela
        for item in tree.get_children():
            tree.delete(item)

        # Inserindo novos dados na tabela
        for row in rows:
            # Substitui valores vazios por uma string vazia
            row = ["" if value is None or value == "" else value for value in row]
            tree.insert('', 'end', values=row)

        # Atualiza o contador de registros
        total_registros.config(text=f"Total de Registros: {len(rows)}")

    except Exception as e:
        messagebox.showerror("Erro na Consulta", f"Erro ao buscar dados: {e}")

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# Função para exportar para Excel com largura fixa, formato de tabela listrada e conteúdo centralizado
def exportar_excel():
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            # Obtendo os dados exibidos na tabela
            rows = [tree.item(item)['values'] for item in tree.get_children()]

            # Convertendo o campo "processo" para string
            for i, row in enumerate(rows):
                rows[i] = list(row)
                rows[i][5] = str(rows[i][5])  # Converter "processo" para string

            # Criando as colunas e os dados
            columns = ['CODIGO', 'DATA', 'GRUPOPESQUISA', 'COD_RELACIONAL', 'NOME', 'PROCESSO', 'PAGINA', 'DIARIO', 'OBS', 'TEXTO', 'STATUS', 'IMPORTADO', 'COD_INTERNO', 'DATA_INCLUSAO', 'ID_ASSUNTO']
            
            # Criando um arquivo Excel com openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Dados Exportados"

            # Adicionando os cabeçalhos
            ws.append(columns)

            # Adicionando os dados
            for row in rows:
                ws.append(row)

            # Ajustando as larguras das colunas conforme especificado
            column_widths = {
                'CODIGO': 17.00,
                'DATA': 17.00,
                'GRUPOPESQUISA': 37.00,
                'COD_RELACIONAL': 23.00,
                'NOME': 13.00,
                'PROCESSO': 31.00,
                'PAGINA': 13.00,
                'DIARIO': 57.00,
                'OBS': 40.00,
                'TEXTO': 40.00,
                'STATUS': 12.00,
                'IMPORTADO': 16.00,
                'COD_INTERNO': 18.00,
                'DATA_INCLUSAO': 20.00,
                'ID_ASSUNTO': 39.57
            }

            for col_idx, col_name in enumerate(columns, start=1):
                column_letter = get_column_letter(col_idx)
                if col_name in column_widths:
                    ws.column_dimensions[column_letter].width = column_widths[col_name]

            # Centralizando os cabeçalhos
            for cell in ws[1]:  # Linha 1 contém os cabeçalhos
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Centralizando todo o conteúdo das células
            for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1, min_col=1, max_col=len(columns)):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Criando uma tabela
            tab = Table(displayName="TabelaDados", ref=f"A1:{get_column_letter(len(columns))}{len(rows) + 1}")

            # Configurando o estilo da tabela com listras
            style = TableStyleInfo(
                name="TableStyleLight1",  # "Branco, Estilo de Tabela Clara 1"
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,  # Listras alternadas nas linhas
                showColumnStripes=False,  # Sem listras nas colunas
            )
            tab.tableStyleInfo = style

            # Adicionando a tabela à planilha
            ws.add_table(tab)

            # Salvando o arquivo
            wb.save(file_path)
            messagebox.showinfo("Sucesso", "Dados exportados com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro na Exportação", f"Erro ao exportar para Excel: {e}")

# Interface gráfica (Tkinter)
root = tk.Tk()
root.title("Relatório Gerencial de Publicação")

# Filtros
frame_filtros = tk.Frame(root)
frame_filtros.pack(pady=10)

tk.Label(frame_filtros, text="Data Inclusão (Início):").grid(row=0, column=0, padx=5)
filtro_data_inicio = tk.Entry(frame_filtros)
filtro_data_inicio.grid(row=0, column=1, padx=5)

tk.Label(frame_filtros, text="Data Inclusão (Fim):").grid(row=0, column=2, padx=5)
filtro_data_fim = tk.Entry(frame_filtros)
filtro_data_fim.grid(row=0, column=3, padx=5)

tk.Label(frame_filtros, text="Data (Início):").grid(row=1, column=0, padx=5)
filtro_data_inicio_data = tk.Entry(frame_filtros)
filtro_data_inicio_data.grid(row=1, column=1, padx=5)

tk.Label(frame_filtros, text="Data (Fim):").grid(row=1, column=2, padx=5)
filtro_data_fim_data = tk.Entry(frame_filtros)
filtro_data_fim_data.grid(row=1, column=3, padx=5)

# Filtro Processo
tk.Label(frame_filtros, text="Processo:").grid(row=2, column=0, padx=5)
filtro_processo = tk.Entry(frame_filtros)
filtro_processo.grid(row=2, column=1, padx=5)

# Filtro Status (Agora abaixo de Processo)
tk.Label(frame_filtros, text="Status:").grid(row=3, column=0, padx=5)
status_var = tk.StringVar(value="Todos")
status_menu = ttk.Combobox(frame_filtros, textvariable=status_var, values=["Todos", "N", "L", "LD", "Excluídos"])
status_menu.grid(row=3, column=1, padx=5)

tk.Button(frame_filtros, text="Filtrar", command=buscar_dados).grid(row=3, column=3, padx=5)

# Tabela de resultados
frame_tabela = tk.Frame(root, padx=15, pady=10)  # Adiciona margens nas laterais e na parte inferior
frame_tabela.pack(pady=10)

# Scrollbar horizontal
scroll_x = ttk.Scrollbar(frame_tabela, orient="horizontal")
scroll_x.pack(side="bottom", fill="x")

# Scrollbar vertical
scroll_y = ttk.Scrollbar(frame_tabela, orient="vertical")
scroll_y.pack(side="right", fill="y")  # Coloca o scrollbar vertical à direita

# Configurando estilo para o cabeçalho da Treeview
style = ttk.Style()
style.theme_use("clam")  # Trocar para o tema "clam" para garantir que estilos sejam aplicados
style.configure("Treeview.Heading", background="#6E6767", foreground="white", font=("Helvetica", 10, "bold"))

# Criando a Treeview
columns = ['CODIGO', 'DATA', 'GRUPOPESQUISA', 'COD_RELACIONAL', 'NOME', 'PROCESSO', 'PAGINA', 'DIARIO', 'OBS', 'TEXTO', 'STATUS', 'IMPORTADO', 'COD_INTERNO', 'DATA_INCLUSAO', 'ID_ASSUNTO']
tree = ttk.Treeview(frame_tabela, columns=columns, show='headings', height=16, xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
tree.pack()

# Configurando as barras de rolagem
scroll_x.config(command=tree.xview)
scroll_y.config(command=tree.yview)

for col in columns:
    tree.heading(col, text=col)


# Configurando as barras de rolagem
scroll_x.config(command=tree.xview)
scroll_y.config(command=tree.yview)

for col in columns:
    tree.heading(col, text=col)

# Label para exibir o total de registros
total_registros = tk.Label(root, text="Total de Registros: 0")
total_registros.pack(pady=5)

# Botão de exportação
tk.Button(root, text="Exportar para Excel", command=exportar_excel).pack(pady=10)

# Iniciar a interface gráfica
root.mainloop()
